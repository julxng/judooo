#!/usr/bin/env python3
"""Generate Judooo Social Media Plan XLSX from markdown content."""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = "/Users/juliusng/Documents/judooo/Judooo_Social_Media_Plan.xlsx"

# ── Style constants ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style="thin", color="D4D4D8"),
    right=Side(style="thin", color="D4D4D8"),
    top=Side(style="thin", color="D4D4D8"),
    bottom=Side(style="thin", color="D4D4D8"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGNMENT = Alignment(wrap_text=True, vertical="center", horizontal="center")

# Sheet tab colors (hex without #)
TAB_COLORS = {
    "Overview":              "3B82F6",  # blue
    "Content Pillars":       "10B981",  # emerald
    "Posting Schedule":      "F59E0B",  # amber
    "Weekly Calendar":       "EF4444",  # red
    "Monthly Series":        "8B5CF6",  # violet
    "Content Formats":       "EC4899",  # pink
    "Hashtags":              "06B6D4",  # cyan
    "Growth Tactics":        "F97316",  # orange
    "KPIs":                  "22C55E",  # green
    "Sample Posts":          "A855F7",  # purple
    "Production Workflow":   "64748B",  # slate
    "Competitive Landscape": "14B8A6",  # teal
}


def style_header_row(ws, row, num_cols):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, bold=False):
    """Apply data cell styling."""
    cell = ws.cell(row=row, column=col)
    cell.font = BOLD_FONT if bold else DATA_FONT
    cell.alignment = WRAP_ALIGNMENT
    cell.border = THIN_BORDER
    return cell


def write_table(ws, start_row, headers, data, bold_first_col=False):
    """Write a table with headers and data. Returns the next available row."""
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=header)
    style_header_row(ws, start_row, len(headers))

    # Write data
    for row_idx, row_data in enumerate(data, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = style_data_cell(ws, row_idx, col_idx, bold=(bold_first_col and col_idx == 1))
            cell.value = value

    return start_row + 1 + len(data)


def write_section_title(ws, row, title, num_cols=1):
    """Write a section title that spans across columns."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SUBTITLE_FONT
    cell.alignment = Alignment(vertical="center")
    if num_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    return row + 1


def auto_size_columns(ws, min_width=12, max_width=60):
    """Auto-size columns based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                # Take the longest line in multi-line content
                lines = str(cell.value).split("\n")
                cell_max = max(len(line) for line in lines)
                max_len = max(max_len, cell_max)
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def freeze_top_row(ws):
    """Freeze the top row for scrolling."""
    ws.freeze_panes = "A2"


# ── Sheet builders ───────────────────────────────────────────────────────────

def build_overview(wb):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_properties.tabColor = TAB_COLORS["Overview"]

    # Title
    cell = ws.cell(row=1, column=1, value="Judooo Social Media Plan")
    cell.font = TITLE_FONT
    ws.merge_cells("A1:C1")

    ws.cell(row=2, column=1, value="Prepared: March 2026").font = DATA_FONT
    ws.cell(row=3, column=1, value="Platform: judooo.com -- Vietnamese art discovery platform").font = DATA_FONT
    ws.cell(row=4, column=1, value="Markets: Ho Chi Minh City, Hanoi, Da Nang, and emerging art cities across Vietnam").font = DATA_FONT

    # Positioning Statement
    row = 6
    row = write_section_title(ws, row, "Positioning Statement", 3)
    ws.cell(row=row, column=1, value=(
        "Judooo is the local art companion for Vietnam -- not a gatekeeping institution, "
        "not an influencer account, but a practical tool that makes discovering and attending "
        "art events feel open, easy, and exciting for anyone curious enough to show up."
    )).font = DATA_FONT
    ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # Voice Characteristics
    row = write_section_title(ws, row, "Voice Characteristics", 3)
    voice_headers = ["Trait", "What It Sounds Like", "What It Does NOT Sound Like"]
    voice_data = [
        ["Approachable", "\"This weekend's lineup has something for everyone -- from first-timers to seasoned collectors.\"", "Academic jargon or elitist art-speak"],
        ["Locally grounded", "Naming specific districts, streets, and neighborhoods (Quan 1, Tay Ho, Dong Da)", "Generic \"Vietnam\" or vague geography"],
        ["Informative-first", "Lead with what, where, when, why go -- then the platform plug", "\"Download our app!\" as the headline"],
        ["Quietly confident", "Sharing real data: \"42 exhibitions happening this month in HCMC\"", "Hyperbolic claims or \"the best art platform ever\""],
        ["Bilingual-natural", "Vietnamese as the primary voice, English seamlessly woven in -- not an afterthought", "Stiff, Google-Translate-quality English translations"],
    ]
    row = write_table(ws, row, voice_headers, voice_data, bold_first_col=True)
    row += 1

    # Bilingual Approach
    row = write_section_title(ws, row, "Bilingual Approach", 3)
    bilingual_data = [
        ["Vietnamese first (primary audience, larger reach on Facebook)"],
        ["Line break / divider (use --- or a simple line gap)"],
        ["English second (expat community, international collectors, SEO value)"],
    ]
    for item in bilingual_data:
        ws.cell(row=row, column=1, value="- " + item[0]).font = DATA_FONT
        ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
    row += 1

    ws.cell(row=row, column=1, value=(
        "Vietnamese captions should feel natural and conversational -- written by someone who lives in Vietnam, "
        "not translated from English. English captions should carry the same tone without being a word-for-word translation."
    )).font = DATA_FONT
    ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    # Platform Personalities
    row = write_section_title(ws, row, "Brand Personality on Each Platform", 3)
    platform_headers = ["Platform", "Role", "Description"]
    platform_data = [
        ["Facebook", "Community hub",
         "Informative, event-forward, conversational. Full event listings, discussions, and gallery partnerships. "
         "94% of Vietnamese internet users are on Facebook."],
        ["Instagram", "Visual portfolio",
         "Curated, aesthetically sharp, discovery-driven. The window into the art itself. "
         "Instagram users in Vietnam tend to be more affluent and aesthetically oriented -- ideal for collectors and design-conscious audiences."],
    ]
    row = write_table(ws, row, platform_headers, platform_data, bold_first_col=True)

    auto_size_columns(ws, max_width=70)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    freeze_top_row(ws)


def build_content_pillars(wb):
    ws = wb.create_sheet("Content Pillars")
    ws.sheet_properties.tabColor = TAB_COLORS["Content Pillars"]

    headers = ["Pillar Name", "Vietnamese Name", "Purpose", "Share %", "Content Types", "Why It Works"]
    data = [
        [
            "Event Discovery",
            "Tuan Nay Co Gi",
            "Drive traffic to judooo.com event listings",
            "35%",
            "Weekly event roundups, individual event highlights, \"Last chance\" reminders, new event announcements, free vs. paid event callouts",
            "This is Judooo's core utility. Every event post is a reason to visit the website. The goal is to become the default answer to \"What art events are happening this weekend?\""
        ],
        [
            "Art & Artist Spotlights",
            "Cau Chuyen Nghe Thuat",
            "Build emotional connection and position Judooo as a platform that values artists",
            "25%",
            "Artist profile features, artwork close-ups from marketplace, gallery/space profiles, \"Meet the maker\" posts, collector stories",
            "Art content is inherently visual and shareable. These posts serve both artists and collectors. Every artist spotlight can link to their artworks on judooo.com."
        ],
        [
            "Art Education & Culture",
            "Nghe Thuat Cho Moi Nguoi",
            "Lower the barrier to entry for new art audiences",
            "20%",
            "Gallery visit guides, art terminology (bilingual), event etiquette tips, Vietnam art history bites, art market trends, route planner tutorials",
            "The biggest barrier to art participation in Vietnam is intimidation, not interest. Educational content expands the total addressable audience."
        ],
        [
            "Community & UGC",
            "Cuoi Tuan Nghe Thuat Cua Ban",
            "Build community, generate social proof, and encourage platform engagement",
            "15%",
            "User-submitted event photos, \"Tag us at the gallery\" campaigns, polls and questions, story reposts, community milestones",
            "UGC creates a flywheel -- people who see others attending events are more likely to attend themselves. Gives galleries and artists a reason to mention Judooo organically."
        ],
        [
            "Platform & Product",
            "Xay Dung Cho Nguoi Yeu Nghe Thuat",
            "Drive feature awareness and website traffic",
            "5%",
            "New feature announcements, \"Did you know?\" tips, before/after comparisons, testimonials from galleries and artists",
            "Keep this pillar lean. Nobody follows a social account to see product updates. But occasional well-crafted product posts convert followers into active users."
        ],
    ]
    write_table(ws, 1, headers, data, bold_first_col=True)
    auto_size_columns(ws, max_width=55)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 55
    freeze_top_row(ws)


def build_posting_schedule(wb):
    ws = wb.create_sheet("Posting Schedule")
    ws.sheet_properties.tabColor = TAB_COLORS["Posting Schedule"]

    row = 1
    row = write_section_title(ws, row, "Facebook Posting Frequency", 4)
    fb_headers = ["Content Type", "Frequency", "Rationale"]
    fb_data = [
        ["Feed posts (events, spotlights, education)", "5x per week", "Facebook's algorithm rewards consistent posting. Vietnam's Facebook audience expects regular updates."],
        ["Stories", "3-5x per week", "Stories maintain top-of-feed presence. Use for quick event reminders, polls, countdowns."],
        ["Facebook Events (co-created with galleries)", "As available", "Co-creating Facebook Events with partner galleries drives direct RSVPs and shows up in users' event feeds."],
        ["Reels", "2x per week", "Facebook is pushing Reels hard. Short video content gets disproportionate reach."],
    ]
    row = write_table(ws, row, fb_headers, fb_data, bold_first_col=True)

    cell = ws.cell(row=row, column=1, value="Total: ~5 feed posts + 2 Reels + 3-5 Stories per week")
    cell.font = BOLD_FONT
    cell.alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    row = write_section_title(ws, row, "Instagram Posting Frequency", 4)
    ig_headers = ["Content Type", "Frequency", "Rationale"]
    ig_data = [
        ["Feed posts (carousel, single image)", "4x per week", "Quality over quantity on Instagram. Each post should be visually polished."],
        ["Reels", "3x per week", "Reels are Instagram's highest-reach format. Art content performs exceptionally well."],
        ["Stories", "Daily (5-7x per week)", "Stories are the heartbeat of Instagram engagement. Polls, questions, countdowns, and reposts keep the account active."],
    ]
    row = write_table(ws, row, ig_headers, ig_data, bold_first_col=True)

    cell = ws.cell(row=row, column=1, value="Total: ~4 feed posts + 3 Reels + daily Stories per week")
    cell.font = BOLD_FONT
    cell.alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    row = write_section_title(ws, row, "Optimal Posting Times (Vietnam -- ICT / UTC+7)", 4)
    time_headers = ["Time Slot", "Platform", "Content Type"]
    time_data = [
        ["7:00-8:00 AM", "Facebook", "Morning event roundups, educational content"],
        ["11:30 AM-1:00 PM", "Facebook & Instagram", "Lunch-break content (carousels, artist spotlights)"],
        ["7:00-9:00 PM", "Instagram & Facebook", "Evening posts (Reels, event highlights, weekend previews)"],
        ["Thursday 7 PM", "Both", "Weekend preview posts (highest intent to plan)"],
        ["Saturday 10 AM", "Both", "Weekend event reminders"],
    ]
    row = write_table(ws, row, time_headers, time_data, bold_first_col=True)

    auto_size_columns(ws, max_width=60)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60
    freeze_top_row(ws)


def build_weekly_calendar(wb):
    ws = wb.create_sheet("Weekly Calendar")
    ws.sheet_properties.tabColor = TAB_COLORS["Weekly Calendar"]

    headers = ["Day", "Facebook", "Instagram", "Theme"]
    data = [
        ["Monday", "Event roundup post: \"This Week in Art\" (HCMC + Hanoi)", "Carousel: Top 5 events this week with visuals", "Event Discovery"],
        ["Tuesday", "Artist/gallery spotlight (long caption + images)", "Reel: 30-second artist studio visit or artwork close-up", "Art & Artist Spotlight"],
        ["Wednesday", "Educational post (art term, gallery guide, collecting tip)", "Carousel: \"Art 101\" educational series", "Art Education"],
        ["Thursday", "Weekend preview post: \"Your Art Weekend Starts Here\"", "Reel: Gallery walkthrough or event preview", "Event Discovery"],
        ["Friday", "Community post (poll, question, UGC repost)", "Story series: Interactive polls + event countdowns", "Community & UGC"],
        ["Saturday", "Event reminder + real-time Story updates from events", "Stories: Live coverage from weekend events", "Event Discovery"],
        ["Sunday", "Rest / scheduled Story only", "Single image: \"Weekend highlight\" best artwork or moment", "Reflection"],
    ]
    write_table(ws, 1, headers, data, bold_first_col=True)
    auto_size_columns(ws, max_width=55)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 22
    freeze_top_row(ws)


def build_monthly_series(wb):
    ws = wb.create_sheet("Monthly Series")
    ws.sheet_properties.tabColor = TAB_COLORS["Monthly Series"]

    headers = ["Series Name", "Cadence", "Platform", "Format"]
    data = [
        ["\"Tuan Nay Co Gi\" (What's On This Week)", "Every Monday", "Both", "Carousel (IG), long post (FB)"],
        ["Artist of the Month", "1st week of month", "Both", "Interview carousel (IG), long-form post (FB)"],
        ["Gallery Guide", "2nd week of month", "Both", "Reel walkthrough"],
        ["\"Art 101\" education series", "Every Wednesday", "Both", "Carousel or infographic"],
        ["Monthly Recap", "Last day of month", "Both", "Carousel: top events, new artists, marketplace highlights"],
        ["Collector Q&A", "3rd week of month", "IG Stories", "Story Q&A series"],
    ]
    write_table(ws, 1, headers, data, bold_first_col=True)
    auto_size_columns(ws, max_width=55)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 48
    freeze_top_row(ws)


def build_content_formats(wb):
    ws = wb.create_sheet("Content Formats")
    ws.sheet_properties.tabColor = TAB_COLORS["Content Formats"]

    row = 1
    row = write_section_title(ws, row, "Facebook Formats", 4)
    fb_headers = ["Format", "Use Case", "Specs", "Tips"]
    fb_data = [
        ["Photo post (single image)", "Event announcements, artwork features", "1200x630px or 1080x1080px", "Always include event date, location, and judooo.com link"],
        ["Photo album", "Multi-event roundups, gallery visit recaps", "3-10 images", "First image should work as standalone; caption each photo"],
        ["Link post", "Driving traffic to judooo.com", "Auto-generates preview card", "Write a compelling caption; don't rely on the link preview alone"],
        ["Video/Reel", "Gallery walkthroughs, artist process, event vibes", "Vertical 9:16, 15-60 seconds", "Hook in first 3 seconds; add Vietnamese subtitles"],
        ["Facebook Event", "Co-created with partner galleries", "Event cover 1920x1005px", "Include Judooo as co-host; link to event detail on judooo.com"],
        ["Stories", "Polls, countdowns, quick updates", "1080x1920px", "Use interactive stickers (poll, quiz, countdown to event)"],
        ["Text post", "Community questions, discussion starters", "No image", "Keep under 150 words; ask a clear question"],
    ]
    row = write_table(ws, row, fb_headers, fb_data, bold_first_col=True)
    row += 1

    row = write_section_title(ws, row, "Instagram Formats", 4)
    ig_headers = ["Format", "Use Case", "Specs", "Tips"]
    ig_data = [
        ["Carousel (up to 20 slides)", "Event roundups, artist features, educational series", "1080x1350px (4:5 portrait)", "Slide 1 is the hook. Last slide is the CTA. Use consistent template design."],
        ["Single image", "Standout artwork, gallery interior, event poster", "1080x1350px (4:5) or 1080x1080 (1:1)", "Clean, high-quality photography. Minimal text overlay."],
        ["Reel", "Gallery walkthroughs, artist process, event atmosphere", "Vertical 9:16, 15-90 seconds", "Trending audio (Vietnamese music when possible). Text overlays for silent viewing."],
        ["Stories", "Daily updates, polls, countdowns, reposts, behind-the-scenes", "1080x1920px", "Use location tags. Poll stickers for engagement. Link sticker to judooo.com."],
        ["Story Highlights", "Permanent categorized content", "Custom cover icons", "Categories: \"This Week\" / \"HCMC\" / \"Hanoi\" / \"Artists\" / \"How To\" / \"Marketplace\""],
        ["Guides", "Curated collections", "Existing posts grouped", "\"Best galleries in District 1\" / \"First-time collector guide\""],
    ]
    row = write_table(ws, row, ig_headers, ig_data, bold_first_col=True)
    row += 1

    row = write_section_title(ws, row, "Visual Style Guide", 4)
    style_data = [
        ["Photography", "Natural light, clean composition, emphasize artwork and gallery spaces. Avoid heavy filters."],
        ["Graphics/Templates", "Use Judooo's design system -- black/white/neutral palette with the Inter font family. Let the art provide the color."],
        ["Carousel templates", "Consistent header bar with Judooo branding. White or off-white backgrounds. Art images as hero elements."],
        ["Reel style", "Steady camera work (gimbal recommended). Subtitles in Vietnamese with English below. Minimal transitions -- let the art and spaces speak."],
    ]
    for item in style_data:
        cell = style_data_cell(ws, row, 1, bold=True)
        cell.value = item[0]
        cell2 = style_data_cell(ws, row, 2)
        cell2.value = item[1]
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

    auto_size_columns(ws, max_width=50)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 50
    freeze_top_row(ws)


def build_hashtags(wb):
    ws = wb.create_sheet("Hashtags")
    ws.sheet_properties.tabColor = TAB_COLORS["Hashtags"]

    row = 1
    row = write_section_title(ws, row, "Branded Hashtags (use on every post)", 3)
    branded_headers = ["Hashtag", "Purpose"]
    branded_data = [
        ["#Judooo", "Primary brand hashtag"],
        ["#JudoooArt", "Art-focused brand tag"],
        ["#JudoooEvents", "Event content tag"],
        ["#KhamPhaNgheThuat", "Vietnamese brand campaign: \"Discover Art\""],
    ]
    row = write_table(ws, row, branded_headers, branded_data, bold_first_col=True)
    row += 1

    row = write_section_title(ws, row, "Location Hashtags (rotate based on content)", 3)
    loc_headers = ["HCMC", "Hanoi", "General Vietnam"]
    loc_data = [
        ["#SaigonArt", "#HanoiArt", "#VietnamArt"],
        ["#SaigonGallery", "#HanoiGallery", "#NgheThuat"],
        ["#ArtSaigon", "#ArtHanoi", "#NgheTheatVietNam"],
        ["#SaigonExhibition", "#HanoiExhibition", "#VietnamExhibition"],
        ["#HoChiMinhCity", "#HaNoiCreative", "#VietnamContemporaryArt"],
        ["#District1Art", "#TayHoArt", "#ArtInVietnam"],
        ["#SaigonCreative", "#OldQuarterArt", "#MadeInVietnam"],
    ]
    row = write_table(ws, row, loc_headers, loc_data)
    row += 1

    row = write_section_title(ws, row, "Category Hashtags (use 3-5 per post, relevant to content)", 3)
    cat_headers = ["Event Types", "Art Mediums", "Audience"]
    cat_data = [
        ["#ArtExhibition", "#ContemporaryArt", "#ArtCollector"],
        ["#TrienLam", "#TrainhSonDau", "#ArtLover"],
        ["#ArtWorkshop", "#LacquerArt", "#NguoiYeuNgheThuat"],
        ["#DauGiaNgheThuat", "#SonMai", "#FirstTimeCollector"],
        ["#ArtAuction", "#Photography", "#WeekendArt"],
        ["#ArtTalk", "#StreetArt", "#ArtCommunity"],
        ["#GalleryOpening", "#Sculpture", "#SuuTamNgheThuat"],
        ["#WorkshopNgheThuat", "#PrintMaking", "#CuoiTuanNgheThuat"],
    ]
    row = write_table(ws, row, cat_headers, cat_data)
    row += 1

    row = write_section_title(ws, row, "Hashtag Campaigns (Seasonal)", 4)
    campaign_headers = ["Campaign", "Hashtag", "Timing", "Purpose"]
    campaign_data = [
        ["Tet art season", "#TetArt #NgheeThuatTet", "January-February", "Tet-themed exhibitions and art gifts"],
        ["Gallery Weekend", "#GalleryWeekendVN", "Quarterly", "Coordinated gallery open-house events"],
        ["Art Month", "#ThangNgheThuat", "Annual (e.g., October)", "Month-long campaign amplifying all events"],
    ]
    row = write_table(ws, row, campaign_headers, campaign_data, bold_first_col=True)
    row += 1

    row = write_section_title(ws, row, "Hashtag Usage Rules", 4)
    rules_headers = ["Platform / Format", "Rule"]
    rules_data = [
        ["Instagram feed posts", "15-20 hashtags (mix of branded + location + category). Place in first comment, not caption."],
        ["Instagram Reels", "5-8 hashtags in caption. Reels rely more on algorithm than hashtag discovery."],
        ["Instagram Stories", "1-3 hashtags (use the hashtag sticker or small text)."],
        ["Facebook posts", "3-5 hashtags maximum. Facebook's algorithm does not reward hashtag stuffing. Use them as category labels."],
    ]
    row = write_table(ws, row, rules_headers, rules_data, bold_first_col=True)

    auto_size_columns(ws, max_width=55)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 44
    freeze_top_row(ws)


def build_growth_tactics(wb):
    ws = wb.create_sheet("Growth Tactics")
    ws.sheet_properties.tabColor = TAB_COLORS["Growth Tactics"]

    headers = ["#", "Tactic Name", "Description", "Target / Partners", "Approach"]
    data = [
        [
            1, "Gallery & Space Partnerships",
            "Establish content partnerships with 10-15 key galleries and art spaces in HCMC and Hanoi.",
            "HCMC: The Factory, Galerie Quynh, San Art, Craig Thomas Gallery, Dia Projects\nHanoi: Manzi, Art Vietnam Gallery, VCCA, The Outpost, Nha San Collective, APD",
            "Galleries tag @judooo when posting; Judooo reposts. Co-create Facebook Events. Provide \"Listed on Judooo\" badge. Monthly cross-promotion."
        ],
        [
            2, "\"Listed on Judooo\" Badge",
            "Create a social graphic that galleries and event organizers can add to their promotional posts.",
            "All event organizers and gallery partners",
            "Turns every partner's post into a referral for Judooo -- organic, scalable, and free."
        ],
        [
            3, "Facebook Group Seeding",
            "Participate actively in existing Vietnam art and culture Facebook groups. Add genuine value, not spam.",
            "Vietnam Art Space, Expat groups in HCMC and Hanoi, Photography/creative communities, University art department groups",
            "Share Monday \"What's On\" roundup as a helpful resource. Answer questions with links to judooo.com. Build reputation as most informed voice."
        ],
        [
            4, "Art & Culture Media Collaboration",
            "Build relationships with Hanoi Grapevine, Saigoneer, Vietcetera, and The Hanoian for content cross-promotion.",
            "Hanoi Grapevine, Saigoneer, Vietcetera, The Hanoian",
            "Offer Judooo event data for their weekly roundups. Guest-contribute event roundup posts. Co-host IG Lives or FB Live sessions."
        ],
        [
            5, "Artist Amplification Program",
            "Feature emerging Vietnamese artists on Judooo's social channels to help them reach new audiences.",
            "Emerging Vietnamese artists with marketplace listings",
            "IG carousel (5 slides), FB long-form interview. Artist receives \"Featured on Judooo\" highlight graphic to share. Each feature shared to artist's network."
        ],
        [
            6, "University & Art School Outreach",
            "Connect with art students and faculty at Vietnam's major art schools.",
            "HCMC University of Fine Arts, Vietnam University of Fine Arts (Hanoi), RMIT Vietnam, Hue University of Arts, BUV",
            "List student exhibitions for free. Run \"Student Spotlight\" series. Partner with student organizations for gallery visits using route planner."
        ],
        [
            7, "Event-Day Live Coverage",
            "Attend 2-3 key openings per month and provide real-time IG Story and FB Story coverage.",
            "Major gallery openings and art events",
            "Stories: Quick clips of space, crowd, standout works. Tag gallery, artists, attendees. End each Story series with link sticker to judooo.com."
        ],
        [
            8, "Seasonal Campaigns",
            "Tie campaigns to Vietnam's cultural calendar.",
            "Tet (Jan-Feb), Int'l Women's Day (Mar 8), Vietnam Heritage Day (Nov 23), Mid-Autumn Festival (Sep), Gallery Weekend (Quarterly), Year-End Auction Season (Nov-Dec)",
            "Create themed content series around each season/event. \"Art gifts for Tet\", female artist features, traditional art forms, collector-focused auction content."
        ],
        [
            9, "Instagram Reels Signature Series",
            "Create 3 recurring Reels series that become recognizable.",
            "All audiences",
            "1) \"60 Giay Tai Gallery\" -- 60-second gallery walkthroughs\n2) \"Truoc vs Sau\" -- Before vs After artist process split-screen\n3) \"Ban Da Biet?\" -- Quick Vietnam art history facts"
        ],
        [
            10, "Paid Amplification Strategy",
            "Monthly ad budget allocation (suggested $300-500/month).",
            "Facebook feed ads (40%): 25-45 yr olds in HCMC & Hanoi\nIG Reels ads (30%): 22-40 yr olds, art/design interest\nFB Event promotion (20%): Geo-targeted, 7-day window\nRetargeting (10%): Website visitors who viewed 2+ events",
            "Focus on event awareness + website traffic (FB), reach + follower growth (IG), RSVPs for partner events, and re-engaging judooo.com visitors."
        ],
    ]
    write_table(ws, 1, headers, data, bold_first_col=False)

    # Make column A (number) narrower, bold the tactic name column
    auto_size_columns(ws, max_width=55)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 55
    freeze_top_row(ws)


def build_kpis(wb):
    ws = wb.create_sheet("KPIs")
    ws.sheet_properties.tabColor = TAB_COLORS["KPIs"]

    row = 1
    row = write_section_title(ws, row, "Awareness Metrics", 4)
    awareness_headers = ["KPI", "Target (Month 1-3)", "Target (Month 4-6)", "Target (Month 7-12)"]
    awareness_data = [
        ["Facebook Page followers", "1,000", "3,000", "8,000"],
        ["Instagram followers", "500", "2,000", "6,000"],
        ["Total monthly reach (FB)", "20,000", "60,000", "150,000"],
        ["Total monthly reach (IG)", "10,000", "40,000", "100,000"],
        ["Reels average views", "500", "2,000", "5,000"],
    ]
    row = write_table(ws, row, awareness_headers, awareness_data, bold_first_col=True)
    row += 1

    row = write_section_title(ws, row, "Engagement Metrics", 4)
    engagement_headers = ["KPI", "Target", "Measurement"]
    engagement_data = [
        ["Facebook engagement rate", "3%+ (feed posts)", "(Reactions + Comments + Shares) / Reach"],
        ["Instagram engagement rate", "4%+ (feed posts)", "(Likes + Comments + Saves + Shares) / Reach"],
        ["Instagram Save rate", "2%+ (carousels)", "Saves / Reach -- indicates high-value content"],
        ["Story completion rate", "70%+", "Percentage of viewers who watch all Story frames"],
        ["Comment-to-like ratio", "5%+", "Comments / Likes -- indicates conversation quality"],
        ["Share rate", "1%+", "Shares / Reach -- indicates viral potential"],
    ]
    row = write_table(ws, row, engagement_headers, engagement_data, bold_first_col=True)
    row += 1

    row = write_section_title(ws, row, "Traffic & Conversion Metrics", 4)
    traffic_headers = ["KPI", "Target", "How to Track"]
    traffic_data = [
        ["Website clicks from social", "500/month by Month 3, 2,000/month by Month 6", "UTM parameters + Plausible analytics (already integrated)"],
        ["Event page views from social", "40% of social clicks land on event detail pages", "Plausible custom events"],
        ["New user signups from social", "50/month by Month 3, 200/month by Month 6", "UTM-tagged registration links"],
        ["Marketplace views from social", "Track separately via UTM", "Plausible"],
        ["Route planner usage from social", "Track via UTM", "Plausible custom events"],
    ]
    row = write_table(ws, row, traffic_headers, traffic_data, bold_first_col=True)
    row += 1

    row = write_section_title(ws, row, "Community Metrics", 4)
    community_headers = ["KPI", "Target", "Measurement"]
    community_data = [
        ["UGC posts per month", "10 by Month 3, 30 by Month 6", "Posts tagging @judooo or using #Judooo"],
        ["Gallery partner reposts", "5/month", "Galleries sharing Judooo content or tagging"],
        ["Artist features shared", "80%+ of featured artists share the post", "Track via Stories reposts and mentions"],
        ["DM inquiries", "Track volume and response time (<2 hours)", "Manual log or social inbox tool"],
    ]
    row = write_table(ws, row, community_headers, community_data, bold_first_col=True)
    row += 1

    row = write_section_title(ws, row, "UTM Parameter Convention", 4)
    utm_headers = ["Example URL", "Description"]
    utm_data = [
        ["https://judooo.com/events?utm_source=facebook&utm_medium=social&utm_campaign=weekly_roundup", "Facebook weekly roundup"],
        ["https://judooo.com/events?utm_source=instagram&utm_medium=social&utm_campaign=artist_spotlight", "Instagram artist spotlight"],
        ["https://judooo.com/marketplace?utm_source=instagram&utm_medium=social&utm_campaign=reel_artwork", "Instagram reel artwork"],
    ]
    row = write_table(ws, row, utm_headers, utm_data)

    cell = ws.cell(row=row, column=1, value="Format: utm_source=[platform]&utm_medium=social&utm_campaign=[content_series]")
    cell.font = BOLD_FONT
    cell.alignment = WRAP_ALIGNMENT

    auto_size_columns(ws, max_width=55)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    freeze_top_row(ws)


def build_sample_posts(wb):
    ws = wb.create_sheet("Sample Posts")
    ws.sheet_properties.tabColor = TAB_COLORS["Sample Posts"]

    headers = ["Post #", "Platform", "Format", "Pillar", "Vietnamese Caption", "English Caption", "Hashtags"]

    vie_1 = (
        "TUAN NAY CO GI | 25-31/3/2026\n\n"
        "5 su kien nghe thuat dang chu y tai Sai Gon tuan nay:\n\n"
        "1. \"Layers of Memory\" -- Trien lam ca nhan cua Nguyen Thi Thanh Mai tai The Factory (Q2). Khai mac thu 5, 26/3. Mien phi.\n"
        "2. Workshop son mai cho nguoi moi bat dau -- Manzi Saigon (Q1). Thu 7, 29/3. 450.000d/nguoi.\n"
        "3. Art Talk: \"Thu Thap Nghe Thuat Dau Tien Cua Ban\" -- Dia Projects (Q3). Thu 6, 28/3. Mien phi.\n"
        "4. Dau gia mua xuan tai Chon Auction House -- Chu nhat, 30/3. Xem truoc tu thu 6.\n"
        "5. \"Urban Sketchers Saigon\" -- Ve ky hoa ngoai troi tai Buu dien TP (Q1). Thu 7, 29/3. Mien phi.\n\n"
        "Xem tat ca su kien + luu vao danh sach tai judooo.com/events"
    )
    en_1 = (
        "WHAT'S ON | 25-31 March 2026\n\n"
        "5 art events worth your time in Saigon this week:\n\n"
        "1. \"Layers of Memory\" -- Solo exhibition by Nguyen Thi Thanh Mai at The Factory (D2). Opening Thursday 26/3. Free entry.\n"
        "2. Lacquer art workshop for beginners -- Manzi Saigon (D1). Saturday 29/3. 450,000 VND/person.\n"
        "3. Art Talk: \"Your First Art Collection\" -- Dia Projects (D3). Friday 28/3. Free.\n"
        "4. Spring auction at Chon Auction House -- Sunday 30/3. Preview from Friday.\n"
        "5. \"Urban Sketchers Saigon\" -- Outdoor sketching at Central Post Office (D1). Saturday 29/3. Free.\n\n"
        "Browse all events + save your favorites at judooo.com/events"
    )
    hashtags_1 = "#Judooo #KhamPhaNgheThuat #SaigonArt #TrienLam #ArtSaigon #WeekendArt"

    vie_2 = (
        "Nghe si thang nay: Tran Duc Minh (@ducminh.studio)\n\n"
        "Tu xuong ve nho tai Binh Thanh, Minh tao ra nhung tac pham son dau lay cam hung tu nhip song thuong nhat cua Sai Gon "
        "-- tu nhung quan ca phe goc pho den anh nang chieu tren nhung con hem.\n\n"
        "\"Toi khong ve nhung gi dep nhat. Toi ve nhung gi that nhat.\"\n\n"
        "3 tac pham cua Minh hien dang co mat tren Judooo Marketplace."
    )
    en_2 = (
        "Artist of the Month: Tran Duc Minh (@ducminh.studio)\n\n"
        "From a small studio in Binh Thanh, Minh creates oil paintings inspired by the everyday rhythms of Saigon "
        "-- from corner cafes to the afternoon light falling through alleyways.\n\n"
        "\"I don't paint what's most beautiful. I paint what's most real.\"\n\n"
        "3 of Minh's works are now available on the Judooo Marketplace.\n\nLink in bio."
    )
    hashtags_2 = "#Judooo #JudoooArt #SaigonArt #VietnamArt #ContemporaryArt #NgheThuat #OilPainting #TrainhSonDau #ArtistSpotlight #VietnamContemporaryArt #ArtCollector #SaigonCreative #BinhThanh #MadeInVietnam #ArtInVietnam"

    vie_3 = (
        "Son mai -- nghe thuat son mai truyen thong Viet Nam -- khong chi la ve. "
        "Do la hang thang kien nhan, tu lop nay sang lop khac.\n\n"
        "Ban da bao gio xem nghe nhan lam son mai chua? Binh luan ben duoi!"
    )
    en_3 = (
        "Vietnamese lacquer art isn't just painting. It's months of patience -- "
        "layer after layer, polished to a glow that no other medium can match.\n\n"
        "Have you ever watched a lacquer artist at work? Tell us below."
    )
    hashtags_3 = "#Judooo #SonMai #LacquerArt #VietnamArt #NgheThuat #ArtEducation #TraditionalArt #VietnamCulture #BanDaBiet #KhamPhaNgheThuat"

    vie_4 = (
        "Hoi nhanh cho dan yeu nghe thuat Sai Gon:\n\n"
        "Neu chi co 1 buoi chieu cuoi tuan, ban se chon:\n"
        "A) Trien lam tranh duong pho tai 3A (Q3)\n"
        "B) Workshop gom thu cong tai The Factory (Q2)\n"
        "C) Dau gia nghe thuat tai Chon (Q1)\n"
        "D) Ve ky hoa ngoai troi tai Ben Thanh\n\n"
        "Binh luan A, B, C, hoac D -- va noi cho minh biet tai sao!\n\n"
        "Tat ca 4 su kien deu co tren judooo.com/events"
    )
    en_4 = (
        "Quick poll for Saigon art lovers:\n\n"
        "If you only had one Saturday afternoon, which would you pick?\n"
        "A) Street art exhibition at 3A (D3)\n"
        "B) Handmade ceramics workshop at The Factory (D2)\n"
        "C) Art auction at Chon (D1)\n"
        "D) Outdoor sketching session at Ben Thanh\n\n"
        "Comment A, B, C, or D -- and tell us why!\n\n"
        "All 4 events listed on judooo.com/events"
    )
    hashtags_4 = "#Judooo #SaigonArt #WeekendArt #CuoiTuanNgheThuat #ArtCommunity"

    vie_5 = (
        "3 trien lam. 1 buoi chieu. 0 phut lang phi.\n\n"
        "Route Planner tren Judooo giup ban sap xep lo trinh tham quan gallery sao cho hop ly nhat "
        "-- chi can luu cac su kien ban thich, va de Judooo toi uu hoa duong di.\n\n"
        "Thu ngay: judooo.com/route-planner"
    )
    en_5 = (
        "3 exhibitions. 1 afternoon. 0 wasted minutes.\n\n"
        "The Route Planner on Judooo helps you map the most efficient gallery-hopping route "
        "-- just save the events you want, and let Judooo optimize the path.\n\n"
        "Try it now: judooo.com/route-planner"
    )
    hashtags_5 = "#Judooo #JudoooEvents #GalleryHopping #ArtWeekend #SaigonArt #RoutePlanner #CuoiTuanNgheThuat"

    data = [
        [1, "Facebook", "Photo album (5 images)", "Event Discovery", vie_1, en_1, hashtags_1],
        [2, "Instagram", "6-slide carousel (1080x1350px)", "Art & Artist Spotlight", vie_2, en_2, hashtags_2],
        [3, "Instagram + Facebook", "Reel (30-45 seconds)", "Art Education", vie_3, en_3, hashtags_3],
        [4, "Facebook", "Text post with single image", "Community & UGC", vie_4, en_4, hashtags_4],
        [5, "Instagram", "4-slide carousel (1080x1350px)", "Platform & Product", vie_5, en_5, hashtags_5],
    ]
    write_table(ws, 1, headers, data, bold_first_col=True)

    # Set row heights for long content
    for r in range(2, 7):
        ws.row_dimensions[r].height = 200

    auto_size_columns(ws, max_width=60)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 55
    ws.column_dimensions["G"].width = 50
    freeze_top_row(ws)


def build_production_workflow(wb):
    ws = wb.create_sheet("Production Workflow")
    ws.sheet_properties.tabColor = TAB_COLORS["Production Workflow"]

    row = 1
    row = write_section_title(ws, row, "Weekly Workflow", 3)
    workflow_headers = ["Day", "Task", "Owner"]
    workflow_data = [
        ["Friday (prev week)", "Draft next week's content calendar. Identify events from judooo.com listings.", "Social Media Manager"],
        ["Saturday-Sunday", "Attend 1-2 events for photo/video content. Capture Stories.", "Social Media Manager / Content Creator"],
        ["Monday morning", "Finalize weekly roundup post. Schedule all week's content.", "Social Media Manager"],
        ["Tuesday-Thursday", "Publish scheduled posts. Monitor comments and DMs. Engage with partner accounts.", "Social Media Manager"],
        ["Friday", "Review week's analytics. Note top-performing content. Adjust next week's plan.", "Social Media Manager"],
    ]
    row = write_table(ws, row, workflow_headers, workflow_data, bold_first_col=True)
    row += 1

    row = write_section_title(ws, row, "Tools Recommended", 3)
    tools_headers = ["Tool", "Purpose"]
    tools_data = [
        ["Meta Business Suite", "Scheduling and analytics for both Facebook and Instagram"],
        ["Canva (or Figma)", "Carousel templates, Story templates, graphic creation"],
        ["CapCut", "Reel editing with Vietnamese subtitle support"],
        ["Plausible (already integrated)", "Website traffic tracking from social with UTM parameters"],
        ["Google Sheets", "Content calendar, KPI tracking, hashtag library"],
        ["Notion", "Editorial planning, partner relationship tracking, content archive"],
    ]
    row = write_table(ws, row, tools_headers, tools_data, bold_first_col=True)

    auto_size_columns(ws, max_width=60)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 34
    freeze_top_row(ws)


def build_competitive_landscape(wb):
    ws = wb.create_sheet("Competitive Landscape")
    ws.sheet_properties.tabColor = TAB_COLORS["Competitive Landscape"]

    headers = ["Account / Platform", "What They Do", "Judooo's Differentiation"]
    data = [
        [
            "Hanoi Grapevine (@grapevine.hanoi)",
            "Bilingual arts & culture event listings for Hanoi. ~6K IG followers.",
            "Judooo covers all of Vietnam (not just Hanoi) and adds interactive features (save, plan routes, marketplace)."
        ],
        [
            "Saigoneer (@saigoneer)",
            "Broad culture/lifestyle media for HCMC. Art is one of many topics.",
            "Judooo is art-only and deeper -- not a media outlet but a utility platform."
        ],
        [
            "The Factory (@thefaboratory)",
            "Single gallery's own account.",
            "Judooo aggregates ALL galleries and events -- not tied to one space."
        ],
        [
            "Vietcetera (@vietcetera)",
            "Vietnamese culture media (English). Art is occasional.",
            "Judooo is art-specific with a functional web tool behind the content."
        ],
        [
            "Individual artist accounts",
            "Artists promote their own work.",
            "Judooo amplifies artists AND gives them marketplace + event submission tools."
        ],
    ]
    row = write_table(ws, 1, headers, data, bold_first_col=True)
    row += 1

    cell = ws.cell(row=row, column=1, value=(
        "Judooo's unique position: No other social account in Vietnam combines art event aggregation, "
        "marketplace access, and a functional web tool. Every post can link to something actionable on "
        "judooo.com -- not just content for content's sake."
    ))
    cell.font = BOLD_FONT
    cell.alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

    auto_size_columns(ws, max_width=60)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 58
    freeze_top_row(ws)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()

    build_overview(wb)
    build_content_pillars(wb)
    build_posting_schedule(wb)
    build_weekly_calendar(wb)
    build_monthly_series(wb)
    build_content_formats(wb)
    build_hashtags(wb)
    build_growth_tactics(wb)
    build_kpis(wb)
    build_sample_posts(wb)
    build_production_workflow(wb)
    build_competitive_landscape(wb)

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
